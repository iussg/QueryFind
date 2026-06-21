import pandas as pd
import io
from datetime import datetime
import re

class ExportManager:
    def to_csv(self, df: pd.DataFrame) -> bytes:
        return df.to_csv(index=False).encode('utf-8')

    def to_excel(self, df: pd.DataFrame, question: str, sql: str) -> bytes:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Results
            df.to_excel(writer, sheet_name='Results', index=False, startrow=0)
            workbook = writer.book
            worksheet = writer.sheets['Results']

            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            # Header styling
            header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

            for col_idx, col in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Alternate row colors
            for row_idx in range(2, len(df) + 2):
                for col_idx in range(1, len(df.columns) + 1):
                    if row_idx % 2 == 0:
                        worksheet.cell(row=row_idx, column=col_idx).fill = alt_fill

            # Auto-fit columns
            for col_idx, col in enumerate(df.columns, 1):
                max_len = max(
                    len(str(col)),
                    df[col].astype(str).str.len().max() if len(df) > 0 else 0
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

            # Sheet 2: Query Info
            info_df = pd.DataFrame([
                ['Question Asked', question],
                ['SQL Generated', sql],
                ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Total Rows', len(df)]
            ], columns=['Field', 'Value'])
            info_df.to_excel(writer, sheet_name='Query Info', index=False)

            # Style query info sheet
            info_ws = writer.sheets['Query Info']
            for col_idx in [1, 2]:
                cell = info_ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
            info_ws.column_dimensions['A'].width = 20
            info_ws.column_dimensions['B'].width = 80

        return output.getvalue()

    def get_filename(self, question: str) -> str:
        clean = re.sub(r'[^a-zA-Z0-9\s]', '', question.lower())
        clean = re.sub(r'\s+', '_', clean.strip())
        return f"querymind_{clean[:50]}"